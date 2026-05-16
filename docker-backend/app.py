# -*- coding: utf-8 -*-
"""
记工App - Flask主应用
=====================

功能模块:
- 用户认证: 注册、登录、登出
- 页面路由: 签到、统计、设置、系统、回收站页面
- 记录管理: 增删改查记工记录
- 回收站: 软删除、恢复、永久删除
- 统计分析: 标准工、加班、饭补、地点统计、工资计算
- 备份管理: 手动备份、自动备份、备份清理
- 数据导出: JSON/Excel多格式导出
- 数据对比: 本月vs上月、同比分析
- 在线更新: Web上传ZIP自动更新

作者: 小扣子
版本: 1.19.0
"""

VERSION = 'v2.1.9.12'

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
import os
import json
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO
from models import (
    init_db, create_user, authenticate_user, get_user_by_id,
    get_all_settings, update_setting, update_settings,
    add_work_record, update_work_record, delete_work_record,
    get_work_records, get_records_by_date_range, get_records_by_year,
    get_records_by_month, get_statistics, clear_all_records,
    get_backup_list, add_backup_record, get_data_hash,
    save_cloud_config, get_cloud_config, delete_cloud_config,
    get_trash_records, restore_work_record, permanent_delete_work_record,
    clean_expired_trash, empty_trash,
    check_duplicate_record, get_monthly_comparison, get_monthly_trend,
    get_current_month_progress,
    # v1.13.0新增
    get_calendar_data, get_yearly_report, import_work_records,
    get_missed_days, validate_hours, get_favorite_locations,
    # v1.16.4新增
    get_work_record_by_id, delete_related_overtime,
    # v1.19.0新增
    get_available_months
)

# ============================================================================
# Flask应用初始化
# ============================================================================

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'work-record-app-secret-key-2026')

# 配置路径
app.config['DATABASE'] = os.environ.get(
    'DATABASE_PATH', 
    os.path.join(os.path.dirname(__file__), 'data', 'work_records.db')
)
app.config['BACKUP_DIR'] = os.environ.get(
    'BACKUP_DIR', 
    os.path.join(os.path.dirname(__file__), 'backups')
)

# 确保必要目录存在
os.makedirs(os.path.dirname(app.config['DATABASE']), exist_ok=True)
os.makedirs(app.config['BACKUP_DIR'], exist_ok=True)

# 初始化数据库
init_db()


# ============================================================================
# 登录验证装饰器
# ============================================================================

def login_required(f):
    """
    登录验证装饰器
    
    支持两种认证方式：
    1. Session认证：Web端通过Cookie登录
    2. Basic Auth认证：App端通过Authorization头登录
    
    验证通过后，当前用户信息会注入到视图函数参数中
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # 先检查session
        if 'user_id' in session:
            return f(*args, **kwargs)
        
        # 再检查Basic Auth（App端使用）
        auth = request.authorization
        if auth and auth.username and auth.password:
            success, message, user = authenticate_user(auth.username, auth.password)
            if success:
                # 将用户信息注入session（不持久化）
                session['user_id'] = user['id']
                session['username'] = user['username']
                return f(*args, **kwargs)
            else:
                # Basic Auth认证失败
                if request.is_json or request.headers.get('Accept') == 'application/json':
                    return jsonify({'success': False, 'message': '认证失败'}), 401
                return redirect(url_for('login'))
        
        # 都没有认证
        if request.is_json or request.headers.get('Accept') == 'application/json':
            return jsonify({'success': False, 'message': '请先登录', 'need_login': True}), 401
        return redirect(url_for('login'))
    return decorated_function


def get_current_user():
    """
    获取当前登录用户
    
    返回:
        dict: 用户信息字典，未登录返回None
    """
    if 'user_id' in session:
        return get_user_by_id(session['user_id'])
    return None


# ============================================================================
# App云同步API（无需登录认证）
# ============================================================================

@app.route('/api/health')
def api_health():
    """
    健康检查端点，供App连接测试使用
    
    返回:
        success: 连接成功标志
        status: 服务状态
    """
    return jsonify({'success': True, 'status': 'ok', 'version': VERSION})


@app.route('/api/sync/version')
def api_sync_version():
    """
    获取同步版本号（供App检查更新使用）
    
    返回:
        success: 成功标志
        version: 版本号字符串
    """
    return jsonify({'success': True, 'version': VERSION})


# ============================================================================
# 用户认证路由
# ============================================================================

@app.route('/offline.html')
def offline():
    """离线页面"""
    return render_template('offline.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    登录页面
    
    GET: 显示登录表单
    POST: 处理登录请求
    """
    # 已登录则跳转到首页
    if 'user_id' in session:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        if not username or not password:
            if request.is_json:
                return jsonify({'success': False, 'message': '请输入用户名和密码'})
            return render_template('login.html', error='请输入用户名和密码')
        
        success, message, user = authenticate_user(username, password)
        
        if success:
            session['user_id'] = user['id']
            session['username'] = user['username']
            if request.is_json:
                return jsonify({'success': True, 'message': '登录成功'})
            return redirect(url_for('index'))
        else:
            if request.is_json:
                return jsonify({'success': False, 'message': message})
            return render_template('login.html', error=message, username=username)
    
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    """
    注册页面
    
    GET: 显示注册表单
    POST: 处理注册请求
    """
    # 已登录则跳转到首页
    if 'user_id' in session:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        username = data.get('username', '').strip()
        password = data.get('password', '')
        confirm_password = data.get('confirm_password', '')
        
        # 验证
        if not username or not password:
            if request.is_json:
                return jsonify({'success': False, 'message': '请输入用户名和密码'})
            return render_template('login.html', error='请输入用户名和密码', show_register=True)
        
        if len(username) < 2 or len(username) > 20:
            if request.is_json:
                return jsonify({'success': False, 'message': '用户名长度应为2-20个字符'})
            return render_template('login.html', error='用户名长度应为2-20个字符', show_register=True)
        
        if len(password) < 4:
            if request.is_json:
                return jsonify({'success': False, 'message': '密码长度至少4个字符'})
            return render_template('login.html', error='密码长度至少4个字符', show_register=True)
        
        if password != confirm_password:
            if request.is_json:
                return jsonify({'success': False, 'message': '两次密码不一致'})
            return render_template('login.html', error='两次密码不一致', show_register=True, username=username)
        
        success, message, user_id = create_user(username, password)
        
        if success:
            # 注册成功后自动登录
            session['user_id'] = user_id
            session['username'] = username
            if request.is_json:
                return jsonify({'success': True, 'message': '注册成功'})
            return redirect(url_for('index'))
        else:
            if request.is_json:
                return jsonify({'success': False, 'message': message})
            return render_template('login.html', error=message, show_register=True, username=username)
    
    return render_template('login.html', show_register=True)


@app.route('/logout')
def logout():
    """
    登出
    
    清除session并跳转到登录页
    """
    session.clear()
    return redirect(url_for('login'))


# ============================================================================
# 页面路由
# ============================================================================

@app.route('/')
@login_required
def index():
    """
    签到主页
    
    功能:
    - 一键记一天（标准工）
    - 手动记工（输入上下班时间）
    - 加班记录（输入加班时长）
    - 工时进度条显示
    - 近期记录列表
    """
    return render_template('index.html', user=get_current_user())


@app.route('/stats')
@login_required
def stats():
    """
    统计页面
    
    功能:
    - 多维度筛选（本月/本年/自定义日期）
    - 标准工统计
    - 加班折算
    - 饭补计算
    - 工资统计
    - 图表可视化（工时趋势、地点占比）
    - 数据对比分析
    """
    return render_template('stats.html', user=get_current_user())


@app.route('/settings')
@login_required
def settings():
    """
    自定义设置页面
    
    功能:
    - 工时规则设置（每日工时、加班折算率、饭补金额）
    - 日工资标准设置
    - 月工时目标设置
    - 自动备份设置（间隔天数、保留份数）
    - 手动备份
    """
    return render_template('settings.html', user=get_current_user())


@app.route('/system')
@login_required
def system():
    """
    系统页面
    
    功能:
    - 数据导出（JSON/Excel格式）
    - 字体大小调节
    - 主题切换（浅色/深色）
    - 清空数据
    - 检查更新
    """
    return render_template('system.html', user=get_current_user())


@app.route('/trash')
@login_required
def trash():
    """
    回收站页面
    
    功能:
    - 查看已删除记录
    - 恢复记录
    - 永久删除
    - 清空回收站
    """
    # 自动清理过期数据
    settings = get_all_settings(session['user_id'])
    retention_days = int(settings.get('trash_retention', 30))
    clean_expired_trash(session['user_id'], retention_days)
    
    return render_template('trash.html', user=get_current_user())


@app.route('/calendar')
@login_required
def calendar():
    """
    日历页面（v1.13.0新增）
    
    功能:
    - 月历视图展示记工记录
    - 点击日期查看详情
    - 左右滑动切换月份
    """
    return render_template('calendar.html', user=get_current_user())


# ============================================================================
# 记录管理API
# ============================================================================

@app.route('/api/records', methods=['GET'])
@login_required
def api_get_records():
    """
    获取记工记录列表
    
    参数:
    - limit: 返回记录数量，默认50条
    """
    limit = request.args.get('limit', 50, type=int)
    records = get_work_records(session['user_id'], limit=limit)
    return jsonify({'success': True, 'data': records})


@app.route('/api/records', methods=['POST'])
@login_required
def api_add_record():
    """
    添加记工记录
    
    请求体:
    - record_type: 记录类型 (standard/manual/overtime)
    - work_date: 工作日期 (YYYY-MM-DD)
    - location: 工作地点（必填）
    - start_time: 开始时间（可选）
    - end_time: 结束时间（可选）
    - morning_end_time: 上午下班时间（可选）
    - afternoon_start_time: 下午上班时间（可选）
    - hours: 工作时长（小时）
    - force_add: 是否强制添加（用于重复打卡确认后）
    - auto_split_overtime: 是否自动拆分加班（标准工或手动折算超过标准工时时生效）
    """
    data = request.get_json()
    
    record_type = data.get('record_type')
    work_date = data.get('work_date')
    location = data.get('location')
    start_time = data.get('start_time')
    end_time = data.get('end_time')
    morning_end_time = data.get('morning_end_time')
    afternoon_start_time = data.get('afternoon_start_time')
    hours = data.get('hours')
    force_add = data.get('force_add', False)
    auto_split_overtime = data.get('auto_split_overtime', False)
    remark = data.get('remark', '')
    raw_meal_subsidy = float(data.get('meal_subsidy', 0)) if data.get('meal_subsidy') else 0
    # 业务规则强制执行：加班没有饭补，标准工必须有饭补，手动折算自由选择
    settings = get_all_settings(session['user_id'])
    meal_subsidy_standard = float(settings.get('meal_subsidy', 30))
    if record_type == 'overtime':
        meal_subsidy = 0
    elif record_type == 'standard':
        meal_subsidy = meal_subsidy_standard
    else:
        meal_subsidy = raw_meal_subsidy
    
    if not all([record_type, work_date, location]):
        return jsonify({'success': False, 'message': '缺少必填字段'})
    
    # 检查重复打卡
    if not force_add:
        duplicate = check_duplicate_record(session['user_id'], work_date, record_type)
        if duplicate['has_duplicate']:
            return jsonify({
                'success': False,
                'message': '检测到该日期已有记工记录，是否继续添加？',
                'duplicate': True,
                'existing_records': duplicate['records']
            })
    
    try:
        # 自动拆分加班：标准工或手动折算超过标准工时时，自动拆分为标准工+加班
        if auto_split_overtime and record_type in ('manual', 'standard') and hours:
            daily_hours = float(settings.get('daily_hours', 9))
            
            if float(hours) > daily_hours:
                overtime_hours = float(hours) - daily_hours
                
                # 创建标准工记录（按标准工时）
                record_id = add_work_record(
                    user_id=session['user_id'],
                    record_type='standard',
                    work_date=work_date,
                    location=location,
                    start_time=start_time,
                    end_time=end_time,
                    morning_end_time=morning_end_time,
                    afternoon_start_time=afternoon_start_time,
                    hours=daily_hours,
                    remark=remark,
                    meal_subsidy=meal_subsidy
                )
                
                # 创建加班记录
                add_work_record(
                    user_id=session['user_id'],
                    record_type='overtime',
                    work_date=work_date,
                    location=location,
                    start_time=None,
                    end_time=None,
                    morning_end_time=None,
                    afternoon_start_time=None,
                    hours=overtime_hours,
                    remark='',
                    meal_subsidy=0
                )
                
                return jsonify({
                    'success': True, 
                    'message': f'记录添加成功：标准工{daily_hours}小时 + 加班{overtime_hours:.1f}小时', 
                    'id': record_id
                })
        
        # 普通记录添加
        record_id = add_work_record(
            user_id=session['user_id'],
            record_type=record_type,
            work_date=work_date,
            location=location,
            start_time=start_time,
            end_time=end_time,
            morning_end_time=morning_end_time,
            afternoon_start_time=afternoon_start_time,
            hours=hours,
            remark=remark,
            meal_subsidy=meal_subsidy
        )
        return jsonify({'success': True, 'message': '记录添加成功', 'id': record_id})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/records/check-duplicate', methods=['POST'])
@login_required
def api_check_duplicate():
    """
    检查重复打卡
    
    请求体:
    - work_date: 工作日期 (YYYY-MM-DD)
    - record_type: 记录类型（可选）
    """
    data = request.get_json()
    work_date = data.get('work_date')
    record_type = data.get('record_type')
    
    if not work_date:
        return jsonify({'success': False, 'message': '缺少日期参数'})
    
    result = check_duplicate_record(session['user_id'], work_date, record_type)
    return jsonify({'success': True, 'data': result})


@app.route('/api/records/<int:record_id>', methods=['PUT'])
@login_required
def api_update_record(record_id):
    """
    更新记工记录
    
    路径参数:
    - record_id: 记录ID
    """
    data = request.get_json()
    
    # 业务规则强制执行：加班没有饭补，标准工必须有饭补，手动折算自由选择
    raw_meal_subsidy = float(data.get('meal_subsidy', 0)) if data.get('meal_subsidy') else None
    record_type = data.get('record_type', '')
    if record_type == 'overtime':
        final_meal_subsidy = 0
    elif record_type == 'standard':
        upd_settings = get_all_settings(session['user_id'])
        final_meal_subsidy = float(upd_settings.get('meal_subsidy', 30))
    else:
        final_meal_subsidy = raw_meal_subsidy
    
    try:
        update_work_record(
            user_id=session['user_id'],
            record_id=record_id,
            record_type=data.get('record_type'),
            work_date=data.get('work_date'),
            location=data.get('location'),
            start_time=data.get('start_time'),
            end_time=data.get('end_time'),
            morning_end_time=data.get('morning_end_time'),
            afternoon_start_time=data.get('afternoon_start_time'),
            hours=data.get('hours'),
            remark=data.get('remark', ''),
            meal_subsidy=final_meal_subsidy
        )
        return jsonify({'success': True, 'message': '记录更新成功'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/records/<int:record_id>', methods=['DELETE'])
@login_required
def api_delete_record(record_id):
    """
    删除记工记录
    
    路径参数:
    - record_id: 记录ID
    
    注意：删除标准工时，会同时删除同一天同地点的关联加班记录
    """
    try:
        # 先获取记录信息
        record = get_work_record_by_id(session['user_id'], record_id)
        
        if not record:
            return jsonify({'success': False, 'message': '记录不存在'})
        
        # 删除主记录
        delete_work_record(session['user_id'], record_id)
        
        # 如果是标准工，删除同一天同地点的关联加班记录
        if record['record_type'] == 'standard':
            delete_related_overtime(session['user_id'], record['work_date'], record['location'])
        
        return jsonify({'success': True, 'message': '记录删除成功'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# ============================================================================
# 统计分析API
# ============================================================================

@app.route('/api/statistics', methods=['GET'])
@login_required
def api_get_statistics():
    """
    获取统计数据
    
    参数（三选一）:
    - year: 按年统计
    - month: 按月统计（配合year参数指定年份）
    - start_date + end_date: 按日期范围统计
    """
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if year:
        stats = get_statistics(session['user_id'], year=year)
    elif month:
        # 支持传入year参数指定具体年份，否则使用当前年份
        target_year = year if year else datetime.now().year
        stats = get_statistics(session['user_id'], month={'year': target_year, 'month': month})
    elif start_date and end_date:
        stats = get_statistics(session['user_id'], start_date=start_date, end_date=end_date)
    else:
        stats = get_statistics(session['user_id'])
    
    return jsonify({'success': True, 'data': stats})


@app.route('/api/statistics/comparison', methods=['GET'])
@login_required
def api_get_comparison():
    """
    获取月度对比数据
    
    返回本月、上月、去年同月的统计数据及对比
    """
    data = get_monthly_comparison(session['user_id'])
    return jsonify({'success': True, 'data': data})


@app.route('/api/statistics/trend', methods=['GET'])
@login_required
def api_get_trend():
    """
    获取工时趋势数据
    
    参数:
    - months: 返回月数，默认6
    """
    months = request.args.get('months', 6, type=int)
    trend_data = get_monthly_trend(session['user_id'], months)
    return jsonify({'success': True, 'data': trend_data})


@app.route('/api/statistics/progress', methods=['GET'])
@login_required
def api_get_progress():
    """
    获取当月工时进度
    """
    progress = get_current_month_progress(session['user_id'])
    return jsonify({'success': True, 'data': progress})


@app.route('/api/available-months', methods=['GET'])
@login_required
def api_get_available_months():
    """
    获取用户有记录的所有月份列表
    
    用于月份选择器下拉框
    """
    months = get_available_months(session['user_id'])
    return jsonify({'success': True, 'data': months})


# ============================================================================
# 设置管理API
# ============================================================================

@app.route('/api/settings', methods=['GET'])
@login_required
def api_get_settings():
    """
    获取所有设置
    """
    settings = get_all_settings(session['user_id'])
    return jsonify({'success': True, 'data': settings})


@app.route('/api/settings', methods=['POST'])
@login_required
def api_update_settings():
    """
    批量更新设置
    """
    data = request.get_json()
    try:
        update_settings(session['user_id'], data)
        return jsonify({'success': True, 'message': '设置更新成功'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/settings/<key>', methods=['PUT'])
@login_required
def api_update_single_setting(key):
    """
    更新单个设置
    """
    data = request.get_json()
    value = data.get('value')
    try:
        update_setting(session['user_id'], key, value)
        return jsonify({'success': True, 'message': '设置更新成功'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# ============================================================================
# 备份管理API
# ============================================================================

@app.route('/api/backup', methods=['POST'])
@login_required
def api_create_backup():
    """
    创建手动备份
    """
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'backup_{session["user_id"]}_{timestamp}.json'
        filepath = os.path.join(app.config['BACKUP_DIR'], filename)
        
        records = get_work_records(session['user_id'], limit=100000)
        settings = get_all_settings(session['user_id'])
        
        backup_data = {
            'user_id': session['user_id'],
            'username': session.get('username'),
            'records': records,
            'settings': settings,
            'export_time': datetime.now().isoformat()
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(backup_data, f, ensure_ascii=False, indent=2)
        
        data_hash = get_data_hash(session['user_id'])
        old_files = add_backup_record(session['user_id'], filename, data_hash)
        
        if old_files:
            for old_file in old_files:
                old_path = os.path.join(app.config['BACKUP_DIR'], old_file)
                if os.path.exists(old_path):
                    os.remove(old_path)
        
        return jsonify({'success': True, 'message': '备份创建成功', 'filename': filename})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/backup/auto', methods=['POST'])
@login_required
def api_auto_backup():
    """
    自动备份
    """
    settings = get_all_settings(session['user_id'])
    interval = int(settings.get('backup_interval', 7))
    
    backup_list = get_backup_list(session['user_id'])
    if backup_list:
        last_backup = datetime.fromisoformat(backup_list[0]['backup_date'])
        if (datetime.now() - last_backup).days < interval:
            return jsonify({'success': True, 'message': '未到备份时间', 'skipped': True})
    
    data_hash = get_data_hash(session['user_id'])
    if backup_list and backup_list[0].get('data_hash') == data_hash:
        return jsonify({'success': True, 'message': '数据无变化', 'skipped': True})
    
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'backup_{session["user_id"]}_{timestamp}.json'
        filepath = os.path.join(app.config['BACKUP_DIR'], filename)
        
        records = get_work_records(session['user_id'], limit=100000)
        backup_data = {
            'user_id': session['user_id'],
            'username': session.get('username'),
            'records': records,
            'settings': settings,
            'export_time': datetime.now().isoformat()
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(backup_data, f, ensure_ascii=False, indent=2)
        
        old_files = add_backup_record(session['user_id'], filename, data_hash)
        
        if old_files:
            for old_file in old_files:
                old_path = os.path.join(app.config['BACKUP_DIR'], old_file)
                if os.path.exists(old_path):
                    os.remove(old_path)
        
        return jsonify({'success': True, 'message': '自动备份完成'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/backup/list', methods=['GET'])
@login_required
def api_list_backups():
    """
    列出所有备份记录
    """
    backups = get_backup_list(session['user_id'])
    return jsonify({'success': True, 'data': backups})


# ============================================================================
# 数据导出API
# ============================================================================

@app.route('/api/export', methods=['GET'])
@login_required
def api_export_data():
    """
    导出数据
    
    参数:
    - type: 导出类型 (year/month/range)
    - format: 导出格式 (json/excel)，默认json
    - year: 年份
    - month: 月份
    - start_date: 开始日期
    - end_date: 结束日期
    """
    export_type = request.args.get('type', 'year')
    export_format = request.args.get('format', 'json')
    year = request.args.get('year', datetime.now().year, type=int)
    month = request.args.get('month', datetime.now().month, type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    try:
        if export_type == 'year':
            records = get_records_by_year(session['user_id'], year)
        elif export_type == 'month':
            records = get_records_by_month(session['user_id'], year, month)
        elif export_type == 'range' and start_date and end_date:
            records = get_records_by_date_range(session['user_id'], start_date, end_date)
        else:
            return jsonify({'success': False, 'message': '无效的导出类型'})
        
        settings = get_all_settings(session['user_id'])
        stats = get_statistics(session['user_id'], year=year) if export_type == 'year' else \
                get_statistics(session['user_id'], month={'year': year, 'month': month}) if export_type == 'month' else \
                get_statistics(session['user_id'], start_date=start_date, end_date=end_date)
        
        if export_format == 'excel':
            # Excel导出
            return export_as_excel(records, stats, settings, export_type, year, month, start_date, end_date)
        else:
            # JSON导出（原有功能）
            export_data = {
                'username': session.get('username'),
                'records': records,
                'statistics': stats,
                'settings': settings,
                'export_time': datetime.now().isoformat(),
                'export_type': export_type
            }
            return jsonify({'success': True, 'data': export_data})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


def export_as_excel(records, stats, settings, export_type, year, month, start_date, end_date):
    """
    导出为Excel文件
    
    参数:
    - records: 记录列表
    - stats: 统计信息
    - settings: 设置信息
    - export_type: 导出类型
    - year, month, start_date, end_date: 日期参数
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        
        # 工作记录表
        ws1 = wb.active
        ws1.title = '工作记录'
        
        # 设置样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4a90d9', end_color='4a90d9', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        headers = ['日期', '类型', '地点', '上班时间', '下班时间', '工时', '备注']
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # 写入数据
        for row, record in enumerate(records, 2):
            type_map = {'standard': '标准工', 'manual': '手动记工', 'overtime': '加班'}
            ws1.cell(row=row, column=1, value=record.get('work_date', '')).border = thin_border
            ws1.cell(row=row, column=2, value=type_map.get(record.get('record_type', ''), '')).border = thin_border
            ws1.cell(row=row, column=3, value=record.get('location', '')).border = thin_border
            ws1.cell(row=row, column=4, value=record.get('start_time', '')).border = thin_border
            ws1.cell(row=row, column=5, value=record.get('end_time', '')).border = thin_border
            ws1.cell(row=row, column=6, value=record.get('hours', 0)).border = thin_border
            ws1.cell(row=row, column=7, value='').border = thin_border
        
        # 设置列宽
        for col in range(1, 8):
            ws1.column_dimensions[get_column_letter(col)].width = 15
        
        # 统计汇总表
        ws2 = wb.create_sheet('统计汇总')
        
        # 写入统计信息
        summary_data = [
            ['记工统计汇总', ''],
            ['', ''],
            ['导出时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['统计周期', f"{year}-{month:02d}" if export_type == 'month' else str(year) if export_type == 'year' else f"{start_date} 至 {end_date}"],
            ['', ''],
            ['=== 工作统计 ===', ''],
            ['标准工天数', stats.get('standard_count', 0)],
            ['加班小时数', stats.get('overtime_hours', 0)],
            ['加班折算', stats.get('overtime_standard', 0)],
            ['手动记工', stats.get('manual_hours', 0)],
            ['总标准工', stats.get('total_standard', 0)],
            ['', ''],
            ['=== 补贴统计 ===', ''],
            ['饭补金额', f"¥ {stats.get('meal_subsidy', 0)}"],
            ['', ''],
            ['=== 工资统计 ===', ''],
            ['日工资标准', f"¥ {stats.get('daily_wage', 0)}"],
            ['应发工资', f"¥ {stats.get('total_wage', 0)}"],
        ]
        
        for row, (label, value) in enumerate(summary_data, 1):
            ws2.cell(row=row, column=1, value=label)
            ws2.cell(row=row, column=2, value=value)
        
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 20
        
        # 生成文件名
        if export_type == 'month':
            filename = f"记工数据_{year}{month:02d}.xlsx"
        elif export_type == 'year':
            filename = f"记工数据_{year}.xlsx"
        else:
            filename = f"记工数据_{start_date}_{end_date}.xlsx"
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except ImportError:
        return jsonify({'success': False, 'message': 'Excel导出功能需要安装openpyxl库'})


# ============================================================================
# 系统管理API
# ============================================================================

@app.route('/api/clear', methods=['POST'])
@login_required
def api_clear_data():
    """
    清空所有数据
    
    警告: 此操作不可恢复！
    """
    try:
        clear_all_records(session['user_id'])
        return jsonify({'success': True, 'message': '数据已清空'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# ============================================================================
# 回收站API
# ============================================================================

@app.route('/api/trash', methods=['GET'])
@login_required
def api_get_trash():
    """
    获取回收站记录列表
    """
    limit = request.args.get('limit', 50, type=int)
    records = get_trash_records(session['user_id'], limit=limit)
    
    # 获取保留天数设置
    settings = get_all_settings(session['user_id'])
    retention_days = int(settings.get('trash_retention', 30))
    
    return jsonify({
        'success': True, 
        'data': records,
        'retention_days': retention_days
    })


@app.route('/api/trash/<int:record_id>/restore', methods=['POST'])
@login_required
def api_restore_record(record_id):
    """
    恢复已删除的记录
    """
    try:
        success = restore_work_record(session['user_id'], record_id)
        if success:
            return jsonify({'success': True, 'message': '记录已恢复'})
        else:
            return jsonify({'success': False, 'message': '记录不存在或已恢复'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/trash/<int:record_id>', methods=['DELETE'])
@login_required
def api_permanent_delete_record(record_id):
    """
    永久删除记录
    """
    try:
        success = permanent_delete_work_record(session['user_id'], record_id)
        if success:
            return jsonify({'success': True, 'message': '记录已永久删除'})
        else:
            return jsonify({'success': False, 'message': '记录不存在'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/trash/empty', methods=['POST'])
@login_required
def api_empty_trash():
    """
    清空回收站
    """
    try:
        count = empty_trash(session['user_id'])
        return jsonify({'success': True, 'message': f'已清空 {count} 条记录'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# ============================================================================
# 邮件备份 API
# ============================================================================

@app.route('/api/cloud/config', methods=['POST'])
@login_required
def api_save_email_config():
    """
    保存邮件备份配置
    
    请求体:
    - smtp_server: SMTP服务器地址
    - smtp_port: SMTP端口
    - smtp_user: 发件邮箱
    - smtp_password: 邮箱授权码/密码
    - receive_email: 收件邮箱
    """
    data = request.get_json()
    
    smtp_server = data.get('smtp_server')
    smtp_port = data.get('smtp_port', 465)
    smtp_user = data.get('smtp_user')
    smtp_password = data.get('smtp_password')
    receive_email = data.get('receive_email')
    
    if not all([smtp_server, smtp_user, smtp_password, receive_email]):
        return jsonify({'success': False, 'message': '请填写完整的配置信息'})
    
    try:
        # 保存配置到数据库
        config_data = json.dumps({
            'smtp_server': smtp_server,
            'smtp_port': smtp_port,
            'smtp_user': smtp_user,
            'smtp_password': smtp_password,
            'receive_email': receive_email
        })
        save_cloud_config(session['user_id'], 'email', config_data=config_data)
        
        return jsonify({'success': True, 'message': '配置已保存'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/cloud/status', methods=['GET'])
@login_required
def api_email_status():
    """
    获取邮件配置状态
    """
    try:
        config = get_cloud_config(session['user_id'], 'email')
        if not config or not config.get('config_data'):
            return jsonify({
                'success': True,
                'connected': False,
                'message': '未配置'
            })
        
        config_data = json.loads(config['config_data'])
        
        return jsonify({
            'success': True,
            'connected': True,
            'message': '已配置',
            'config': {
                'smtp_server': config_data.get('smtp_server'),
                'smtp_user': config_data.get('smtp_user'),
                'receive_email': config_data.get('receive_email')
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/cloud/test', methods=['POST'])
@login_required
def api_test_email():
    """
    测试邮件发送
    """
    try:
        config = get_cloud_config(session['user_id'], 'email')
        if not config or not config.get('config_data'):
            return jsonify({'success': False, 'message': '请先配置邮件'})
        
        config_data = json.loads(config['config_data'])
        
        # 发送测试邮件
        send_test_email(config_data)
        
        return jsonify({'success': True, 'message': '测试邮件已发送'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'发送失败: {str(e)}'})


@app.route('/api/cloud/backup', methods=['POST'])
@login_required
def api_send_backup_email():
    """
    发送备份邮件
    """
    try:
        config = get_cloud_config(session['user_id'], 'email')
        if not config or not config.get('config_data'):
            return jsonify({'success': False, 'message': '请先配置邮件'})
        
        config_data = json.loads(config['config_data'])
        
        # 获取数据库文件路径
        db_path = os.environ.get('DATABASE_PATH', '/app/data/work_records.db')
        
        if not os.path.exists(db_path):
            return jsonify({'success': False, 'message': '数据库文件不存在'})
        
        # 发送备份邮件
        send_backup_email(config_data, db_path)
        
        return jsonify({'success': True, 'message': '备份邮件已发送'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'发送失败: {str(e)}'})


def send_test_email(config):
    """
    发送测试邮件
    
    参数:
        config: 邮件配置字典
    """
    msg = MIMEMultipart()
    msg['From'] = config['smtp_user']
    msg['To'] = config['receive_email']
    msg['Subject'] = '【记工本】测试邮件'
    
    body = f"""
您好！

这是一封来自【记工本】的测试邮件。

发送时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

如果您收到此邮件，说明邮件配置成功！

---
记工本 - 记录每一天的辛勤付出
"""
    
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    
    # 发送邮件
    with smtplib.SMTP_SSL(config['smtp_server'], config['smtp_port']) as server:
        server.login(config['smtp_user'], config['smtp_password'])
        server.send_message(msg)


def send_backup_email(config, db_path):
    """
    发送备份邮件
    
    参数:
        config: 邮件配置字典
        db_path: 数据库文件路径
    """
    msg = MIMEMultipart()
    msg['From'] = config['smtp_user']
    msg['To'] = config['receive_email']
    msg['Subject'] = f'【记工本】数据备份 {datetime.now().strftime("%Y%m%d %H:%M")}'
    
    body = f"""
您好！

这是来自【记工本】的自动数据备份。

备份时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

请注意查收附件中的数据库文件。

---
记工本 - 记录每一天的辛勤付出
"""
    
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    
    # 添加附件
    with open(db_path, 'rb') as f:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(f.read())
        encoders.encode_base64(part)
        filename = f'work_records_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db'
        part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(part)
    
    # 发送邮件
    with smtplib.SMTP_SSL(config['smtp_server'], config['smtp_port']) as server:
        server.login(config['smtp_user'], config['smtp_password'])
        server.send_message(msg)


# ============================================================================
# v1.13.0 新增 API
# ============================================================================

@app.route('/api/calendar/<int:year>/<int:month>', methods=['GET'])
@login_required
def api_get_calendar(year, month):
    """
    获取日历数据
    
    参数:
    - year: 年份
    - month: 月份（1-12）
    """
    try:
        calendar_data = get_calendar_data(session['user_id'], year, month)
        return jsonify({'success': True, 'data': calendar_data})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/yearly-report/<int:year>', methods=['GET'])
@login_required
def api_get_yearly_report(year):
    """
    获取年度工作报告
    """
    try:
        report = get_yearly_report(session['user_id'], year)
        return jsonify({'success': True, 'data': report})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/import', methods=['POST'])
@login_required
def api_import_data():
    """
    导入数据
    
    支持三种格式:
    1. application/json: JSON批量导入 {"records": [...]}（供App云同步使用）
    2. multipart/form-data + JSON文件: 上传.json文件
    3. multipart/form-data + Excel文件: 上传.xlsx/.xls文件
    """
    try:
        # 获取用户设置（用于饭补金额和标准工时）
        user_settings = get_all_settings(session['user_id'])
        meal_subsidy_amount = float(user_settings.get('meal_subsidy', 30))
        daily_hours = float(user_settings.get('daily_hours', 9))
        
        # 辅助函数：强制执行饭补业务规则
        def enforce_meal_subsidy(record_type, raw_meal_subsidy):
            """业务规则：标准工必须有饭补，加班没有饭补，手动折算自由选择"""
            if record_type == 'overtime':
                return 0  # 加班没有饭补
            elif record_type == 'standard':
                return meal_subsidy_amount  # 标准工必须有饭补
            else:  # manual
                # 手动折算：如果传了True/有值则给饭补，否则不给
                if isinstance(raw_meal_subsidy, bool):
                    return meal_subsidy_amount if raw_meal_subsidy else 0
                elif isinstance(raw_meal_subsidy, (int, float)):
                    return meal_subsidy_amount if raw_meal_subsidy > 0 else 0
                else:
                    return meal_subsidy_amount  # 默认有饭补
        
        # 辅助函数：处理记录列表
        def process_records(records):
            imported = 0
            duplicates = 0
            for record in records:
                work_date = record.get('work_date') or record.get('date')
                record_type = record.get('record_type', 'standard')
                location = record.get('location', '')
                hours = float(record.get('hours', 8) or 8)
                raw_meal_subsidy = record.get('meal_subsidy', 0)
                
                if not work_date:
                    continue
                
                # 检查是否重复
                duplicate = check_duplicate_record(session['user_id'], work_date, record_type)
                if duplicate['has_duplicate']:
                    duplicates += 1
                    continue
                
                # 强制执行饭补业务规则
                final_meal_subsidy = enforce_meal_subsidy(record_type, raw_meal_subsidy)
                
                # 添加记录
                add_work_record(
                    user_id=session['user_id'],
                    record_type=record_type,
                    work_date=work_date,
                    location=location,
                    start_time=record.get('start_time'),
                    end_time=record.get('end_time'),
                    morning_end_time=record.get('morning_end_time'),
                    afternoon_start_time=record.get('afternoon_start_time'),
                    hours=hours,
                    remark=record.get('remark', ''),
                    meal_subsidy=final_meal_subsidy
                )
                imported += 1
            return imported, duplicates
        
        # 方式1: JSON请求体（App云同步使用）
        if request.is_json:
            data = request.get_json()
            records = data.get('records', [])
            if not records:
                return jsonify({'success': False, 'message': '没有要导入的记录'})
            imported, duplicates = process_records(records)
            return jsonify({
                'success': True,
                'data': {'success_count': imported, 'error_count': duplicates},
                'message': f'成功导入 {imported} 条记录，跳过 {duplicates} 条重复记录'
            })
        
        # 方式2/3: FormData文件上传
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '请选择文件'})
        
        file = request.files['file']
        filename = file.filename or ''
        
        # JSON文件导入
        if filename.endswith('.json'):
            try:
                json_str = file.read().decode('utf-8')
                import json as json_module
                data = json_module.loads(json_str)
                
                # 兼容多种JSON格式
                records = None
                if isinstance(data, list):
                    records = data
                elif isinstance(data, dict):
                    # 尝试多种格式
                    if 'records' in data:
                        records = data['records']
                    elif 'data' in data and isinstance(data['data'], dict) and 'records' in data['data']:
                        records = data['data']['records']
                
                if not records:
                    return jsonify({'success': False, 'message': 'JSON文件中未找到记录数据'})
                
                imported, duplicates = process_records(records)
                return jsonify({
                    'success': True,
                    'data': {'success_count': imported, 'error_count': duplicates},
                    'message': f'成功导入 {imported} 条记录，跳过 {duplicates} 条重复记录'
                })
            except Exception as e:
                return jsonify({'success': False, 'message': f'JSON解析失败：{str(e)}'})
        
        # Excel文件导入
        if not filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': '请上传JSON或Excel文件(.json/.xlsx/.xls)'})
        
        # 读取Excel
        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active
        
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            records.append({
                'record_type': row[1] if len(row) > 1 else 'standard',
                'work_date': str(row[0])[:10] if row[0] else '',
                'location': row[2] if len(row) > 2 else '',
                'start_time': row[3] if len(row) > 3 else None,
                'end_time': row[4] if len(row) > 4 else None,
                'hours': float(row[5]) if len(row) > 5 and row[5] else 8,
                'meal_subsidy': row[6] if len(row) > 6 else 0
            })
        
        # 使用统一的process_records处理（自动执行饭补业务规则）
        imported, duplicates = process_records(records)
        
        return jsonify({
            'success': True,
            'data': {'success_count': imported, 'error_count': duplicates},
            'message': f'成功导入 {imported} 条记录，跳过 {duplicates} 条重复记录'
        })
    except ImportError:
        return jsonify({'success': False, 'message': 'Excel导入功能需要安装openpyxl库'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/missed-days', methods=['GET'])
@login_required
def api_get_missed_days():
    """
    获取漏记日期
    """
    days = request.args.get('days', 7, type=int)
    missed_days = get_missed_days(session['user_id'], days)
    return jsonify({'success': True, 'data': missed_days})


@app.route('/api/validate-hours', methods=['POST'])
@login_required
def api_validate_hours():
    """
    校验工时异常
    """
    data = request.get_json()
    record_type = data.get('record_type', 'manual')
    hours = float(data.get('hours', 0))
    
    result = validate_hours(record_type, hours)
    return jsonify({'success': True, 'data': result})


@app.route('/api/favorite-locations', methods=['GET'])
@login_required
def api_get_favorite_locations():
    """
    获取常用地点
    """
    limit = request.args.get('limit', 5, type=int)
    locations = get_favorite_locations(session['user_id'], limit)
    return jsonify({'success': True, 'data': locations})


@app.route('/api/locations', methods=['GET'])
@login_required
def api_get_all_locations():
    """
    获取所有工作地点（用于地点选择下拉框）
    """
    conn = models.get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT DISTINCT location FROM work_records 
        WHERE user_id = ? AND deleted_at IS NULL
        ORDER BY location
    ''', (session['user_id'],))
    rows = cursor.fetchall()
    conn.close()
    return jsonify({'success': True, 'data': [row['location'] for row in rows]})


@app.route('/api/user/change-password', methods=['POST'])
@login_required
def api_change_password():
    """
    修改密码
    """
    data = request.get_json()
    current_password = data.get('current_password', '')
    new_password = data.get('new_password', '')
    
    if not current_password or not new_password:
        return jsonify({'success': False, 'message': '请填写所有字段'})
    
    if len(new_password) < 4:
        return jsonify({'success': False, 'message': '新密码长度至少4位'})
    
    # 验证当前密码
    user = models.get_user_by_id(session['user_id'])
    if not user:
        return jsonify({'success': False, 'message': '用户不存在'})
    
    # 验证密码
    import hashlib
    hashed_current = hashlib.sha256((current_password + user['salt']).encode()).hexdigest()
    if hashed_current != user['password']:
        return jsonify({'success': False, 'message': '当前密码错误'})
    
    # 更新密码
    new_hash = hashlib.sha256((new_password + user['salt']).encode()).hexdigest()
    conn = models.get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE users SET password = ? WHERE id = ?', (new_hash, session['user_id']))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': '密码修改成功'})


# ============================================================================
# 管理员API
# ============================================================================

@app.route('/api/admin/users', methods=['GET'])
@login_required
def api_admin_list_users():
    """
    获取用户列表（仅admin用户可访问）
    """
    if session.get('username') != 'admin':
        return jsonify({'success': False, 'message': '无权限，只有管理员可以访问'})
    
    conn = models.get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT id, username, created_at FROM users ORDER BY id')
    users = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify({'success': True, 'data': users})


@app.route('/api/admin/reset-password', methods=['POST'])
@login_required
def api_admin_reset_password():
    """
    管理员重置用户密码（仅admin用户可访问）
    """
    if session.get('username') != 'admin':
        return jsonify({'success': False, 'message': '无权限，只有管理员可以访问'})
    
    data = request.get_json()
    user_id = data.get('user_id')
    new_password = data.get('new_password', '123456')
    
    if not user_id:
        return jsonify({'success': False, 'message': '参数不完整'})
    
    if len(new_password) < 4:
        return jsonify({'success': False, 'message': '密码至少4位'})
    
    # 检查用户是否存在
    user = models.get_user_by_id(user_id)
    if not user:
        return jsonify({'success': False, 'message': '用户不存在'})
    
    # 更新密码
    import hashlib
    new_hash = hashlib.sha256((new_password + user['salt']).encode()).hexdigest()
    conn = models.get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE users SET password = ? WHERE id = ?', (new_hash, user_id))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': '密码重置成功'})


# ============================================================================
# 启动入口
# ============================================================================

if __name__ == '__main__':
    # 本地开发模式
    app.run(host='0.0.0.0', port=8080, debug=True)


# ============================================================================
# 在线更新功能
# ============================================================================

@app.route('/api/version')
@login_required
def api_version():
    """获取当前版本号"""
    return jsonify({'success': True, 'version': VERSION})


@app.route('/api/update', methods=['POST'])
@login_required
def api_update():
    """在线更新"""
    import tempfile
    import zipfile
    import shutil
    
    # 验证文件
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未选择文件'})
    
    file = request.files['file']
    if not file.filename.endswith('.zip'):
        return jsonify({'success': False, 'message': '请上传ZIP文件'})
    
    try:
        # 备份数据库（备份到/app/backups目录，确保持久化）
        db_path = os.environ.get('DATABASE_PATH', '/app/data/work_records.db')
        backup_dir = os.environ.get('BACKUP_DIR', '/app/backups')
        os.makedirs(backup_dir, exist_ok=True)
        backup_path = os.path.join(backup_dir, f'auto_backup_{datetime.now().strftime("%Y%m%d%H%M%S")}.db')
        if os.path.exists(db_path):
            shutil.copy2(db_path, backup_path)
        
        # 解压ZIP
        temp_dir = tempfile.mkdtemp()
        zip_path = os.path.join(temp_dir, 'update.zip')
        file.save(zip_path)
        
        with zipfile.ZipFile(zip_path, 'r') as zf:
            zf.extractall(temp_dir)
        
        # 查找应用目录
        app_source = temp_dir
        for root, dirs, files in os.walk(temp_dir):
            if 'app.py' in files:
                app_source = root
                break
        
        # 复制文件
        app_dir = '/app'
        
        for filename in ['app.py', 'models.py', 'requirements.txt']:
            src = os.path.join(app_source, filename)
            if os.path.exists(src):
                shutil.copy2(src, os.path.join(app_dir, filename))
        
        for dirname in ['templates', 'static']:
            src = os.path.join(app_source, dirname)
            if os.path.exists(src):
                dst = os.path.join(app_dir, dirname)
                if os.path.exists(dst):
                    shutil.rmtree(dst)
                shutil.copytree(src, dst)
        
        # 写入重启标记
        with open(os.path.join(app_dir, '.need_restart'), 'w') as f:
            f.write('1')
        
        # 清理临时目录
        shutil.rmtree(temp_dir)
        
        return jsonify({'success': True, 'message': '更新成功，正在重启...'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'更新失败: {str(e)}'})


@app.before_request
def check_restart():
    """检查是否需要重启"""
    restart_marker = '/app/.need_restart'
    if os.path.exists(restart_marker):
        if request.path != '/api/update':
            try:
                os.remove(restart_marker)
            except:
                pass
            import threading
            threading.Timer(1, lambda: os._exit(0)).start()
